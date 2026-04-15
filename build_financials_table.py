"""
Build the Deal Financials table in the 'Financials' tab of the Excel workbook.

Uses direct range references to the '_DealFinancial' sheet (e.g.
`_DealFinancial!$I:$I`) instead of structured-table references like
`cr83c_financials[Amount]`. The structured-reference version used to work
visually but openpyxl does not round-trip the underlying table definition,
so after save the refs would collapse to #REF!. Direct range refs have no
such dependency.

Layout (matches the Power BI visual, minus the $/unit column):

  Adj. GPR                = GPR + Loss to Lease
  Effective Rental Income = Adj. GPR + Vacancy + Concessions + Bad Debt Loss
  Effective Gross Rev.    = Effective Rental Income + Utility Income + Other Income
  Total Controllable      = G&A + Payroll + S&M + Utilities + R&M + Mgmt Fee + Turnover
  Total Operating Exp.    = Total Controllable + Property Taxes + Insurance
  Net Operating Income    = Effective Gross Revenue - Total Operating Expenses
  Cash Flow from Ops      = NOI - CapEx Reserve - Debt Service

% of GPR denominator:
  - Revenue block (rows 7..14)          -> Adj. GPR           (row 7)
  - OpEx / NOI / Below NOI / CashFlow   -> Eff. Gross Revenue (row 14)

Columns:
  B = Label
  C = ACT $
  D = ACT % of GPR
  E = UW  $
  F = UW  % of GPR
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORKBOOK = "Dynamics Workbook_03.23.2026 1.xlsx"
DEAL_REF = "COVER!$C$2"  # active deal filter

# Direct range refs into the _DealFinancial sheet.
DEAL_COL = "_DealFinancial!$A:$A"   # Deal Name
CAT_COL  = "_DealFinancial!$E:$E"   # Category (ACT / UW)
ILI_COL  = "_DealFinancial!$H:$H"   # Internal Line Item
AMT_COL  = "_DealFinancial!$I:$I"   # Amount

# ---------- styling ----------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SECTION_FILL = PatternFill("solid", fgColor="D9D9D9")
SUBTOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
DATA_FILL = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Arial", size=11, bold=True, color="000000")
SUBTOTAL_FONT = Font(name="Arial", size=11, bold=True, color="000000")
DATA_FONT = Font(name="Arial", size=11, bold=False, color="000000")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT_SECTION = Alignment(horizontal="left", vertical="center", indent=1)
LEFT_DATA = Alignment(horizontal="left", vertical="center", indent=2)
RIGHT = Alignment(horizontal="right", vertical="center")

FMT_CURRENCY = '"$"#,##0;[Red]("$"#,##0);"$"0'
FMT_PCT = '0.0%;[Red]-0.0%;0.0%'

# ---------- Layout definition ----------
# Each row: (label, row_type, item / tag)
#   row_type: 'section' | 'data' | 'subtotal' | 'final'
LAYOUT = [
    ("Revenue",                     "section",  None),
    ("Adj. GPR",                    "subtotal", "ADJ_GPR"),
    ("Vacancy",                     "data",     "Vacancy"),
    ("Concessions",                 "data",     "Concessions"),
    ("Bad Debt Loss",               "data",     "Bad Debt Loss"),
    ("Effective Rental Income",     "subtotal", "EFF_RENTAL"),
    ("Utility Income",              "data",     "Utility Income"),
    ("Other Income",                "data",     "Other Income"),
    ("Effective Gross Revenue",     "subtotal", "EFF_GROSS_REV"),
    ("Operating Expenses",          "section",  None),
    ("General & Administrative",    "data",     "General & Administrative"),
    ("Payroll",                     "data",     "Payroll"),
    ("Sales & Marketing",           "data",     "Sales & Marketing"),
    ("Utilities",                   "data",     "Utilities"),
    ("Repairs & Maintenance",       "data",     "Repairs & Maintenance"),
    ("Management Fee",              "data",     "Management Fee"),
    ("Turnover",                    "data",     "Turnover"),
    ("Total Controllable Expenses", "subtotal", "TOTAL_CTRL"),
    ("Property Taxes",              "data",     "Property Taxes"),
    ("Insurance",                   "data",     "Insurance"),
    ("Total Operating Expenses",    "subtotal", "TOTAL_OPEX"),
    ("Net Operating Income",        "final",    "NOI"),
    ("Below NOI",                   "section",  None),
    ("CapEx Reserve",               "data",     "CapEx Reserve"),
    ("Debt Service",                "data",     "Debt Service"),
    ("Cash Flow from Operations",   "final",    "CASHFLOW"),
]

PCT_DENOM_REVENUE_BLOCK = {
    "ADJ_GPR", "Vacancy", "Concessions", "Bad Debt Loss",
    "EFF_RENTAL", "Utility Income", "Other Income", "EFF_GROSS_REV",
}


# ---------- formula builders ----------

def sumifs_item(item, category):
    """SUMIFS for a single Internal Line Item, using direct ranges."""
    return (
        f'=SUMIFS({AMT_COL},'
        f'{DEAL_COL},{DEAL_REF},'
        f'{CAT_COL},"{category}",'
        f'{ILI_COL},"{item}")'
    )


def sumifs_multi(items, category):
    """SUMIFS over a list of Internal Line Items (wrapped in SUM for the array)."""
    arr = "{" + ",".join(f'"{i}"' for i in items) + "}"
    return (
        f'=SUM(SUMIFS({AMT_COL},'
        f'{DEAL_COL},{DEAL_REF},'
        f'{CAT_COL},"{category}",'
        f'{ILI_COL},{arr}))'
    )


def formula_for_dollars(item_tag, category, row_map):
    """Build the $ formula for a given layout row tag."""
    if item_tag == "ADJ_GPR":
        return sumifs_multi(["Gross Potential Rent", "Loss to Lease"], category)
    if item_tag == "EFF_RENTAL":
        r_adj = row_map["ADJ_GPR"]
        r_vac = row_map["Vacancy"]
        r_con = row_map["Concessions"]
        r_bdl = row_map["Bad Debt Loss"]
        col = "{COL}"
        return f"={col}{r_adj}+{col}{r_vac}+{col}{r_con}+{col}{r_bdl}"
    if item_tag == "EFF_GROSS_REV":
        r_er = row_map["EFF_RENTAL"]
        r_ui = row_map["Utility Income"]
        r_oi = row_map["Other Income"]
        col = "{COL}"
        return f"={col}{r_er}+{col}{r_ui}+{col}{r_oi}"
    if item_tag == "TOTAL_CTRL":
        items = ["General & Administrative", "Payroll", "Sales & Marketing",
                 "Utilities", "Repairs & Maintenance", "Management Fee", "Turnover"]
        rows = [row_map[i] for i in items]
        col = "{COL}"
        return "=" + "+".join(f"{col}{r}" for r in rows)
    if item_tag == "TOTAL_OPEX":
        r_tc = row_map["TOTAL_CTRL"]
        r_pt = row_map["Property Taxes"]
        r_in = row_map["Insurance"]
        col = "{COL}"
        return f"={col}{r_tc}+{col}{r_pt}+{col}{r_in}"
    if item_tag == "NOI":
        r_egr = row_map["EFF_GROSS_REV"]
        r_top = row_map["TOTAL_OPEX"]
        col = "{COL}"
        return f"={col}{r_egr}-{col}{r_top}"
    if item_tag == "CASHFLOW":
        r_noi = row_map["NOI"]
        r_cap = row_map["CapEx Reserve"]
        r_ds = row_map["Debt Service"]
        col = "{COL}"
        return f"={col}{r_noi}-{col}{r_cap}-{col}{r_ds}"
    # plain data row — item_tag IS the Internal Line Item value
    return sumifs_item(item_tag, category)


# ---------- sheet clearing ----------

def clear_sheet(ws):
    """
    Wipe all cells, merges, data validations, and column dimensions from the
    Financials sheet. We rebuild it from scratch on every run.
    """
    # Drop every merged range
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))

    # Delete row dimensions so we can re-apply heights cleanly
    ws.row_dimensions.clear()

    # Drop all data validations (covers any legacy date-picker dropdown)
    ws.data_validations.dataValidation = []

    # Wipe cells in a generous box
    max_r = max(ws.max_row, 50)
    max_c = max(ws.max_column, 40)
    for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.font = Font()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = "General"

    # Reset column widths / hidden state across a wide range. openpyxl does
    # not accept None for width, so use the default width as a neutral reset.
    # customWidth is read-only; it's auto-set when width is assigned.
    for c in range(1, max_c + 1):
        letter = get_column_letter(c)
        cd = ws.column_dimensions[letter]
        cd.width = 10.0
        cd.hidden = False


def build_table(ws):
    col_label = "B"
    cols_act = ["C", "D"]  # $, % of GPR
    cols_uw  = ["E", "F"]

    header_r1 = 4  # Category | ACT | UW
    header_r2 = 5  # Section  | $ | % | $ | %
    first_data_row = header_r2 + 1  # 6

    # Pre-compute each tag's row for cross-ref in subtotal / final formulas.
    row_map = {}
    r = first_data_row
    for (label, typ, tag) in LAYOUT:
        if typ != "section":
            row_map[tag] = r
        r += 1

    clear_sheet(ws)

    # ---------- row 2: deal name display ------------------------------------
    ws["B2"] = f"={DEAL_REF}"
    ws["B2"].font = Font(name="Arial", size=14, bold=True, color="1F3864")
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    # ---------- header rows --------------------------------------------------
    ws[f"B{header_r1}"] = "Category"
    ws[f"B{header_r2}"] = "Section"
    ws.merge_cells(f"B{header_r1}:B{header_r2}")

    ws[f"C{header_r1}"] = "ACT"
    ws.merge_cells(f"C{header_r1}:D{header_r1}")
    ws[f"E{header_r1}"] = "UW"
    ws.merge_cells(f"E{header_r1}:F{header_r1}")

    # Sub-headers: $ and % of GPR
    ws[f"C{header_r2}"] = "$"
    ws[f"D{header_r2}"] = "% of GPR"
    ws[f"E{header_r2}"] = "$"
    ws[f"F{header_r2}"] = "% of GPR"

    for col in ["B"] + cols_act + cols_uw:
        for rr in (header_r1, header_r2):
            c = ws[f"{col}{rr}"]
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER
            c.border = BORDER_ALL

    ws.row_dimensions[header_r1].height = 20
    ws.row_dimensions[header_r2].height = 20

    # ---------- body ---------------------------------------------------------
    r = first_data_row
    for (label, typ, tag) in LAYOUT:
        lc = ws[f"{col_label}{r}"]
        lc.value = label
        lc.border = BORDER_ALL

        if typ == "section":
            lc.fill = SECTION_FILL
            lc.font = SECTION_FONT
            lc.alignment = LEFT_SECTION
            for col in cols_act + cols_uw:
                c = ws[f"{col}{r}"]
                c.fill = SECTION_FILL
                c.border = BORDER_ALL
            r += 1
            continue

        # Choose styles for data / subtotal / final rows
        if typ in ("subtotal", "final"):
            fill = SUBTOTAL_FILL
            font = SUBTOTAL_FONT
            lc.alignment = LEFT_SECTION
        else:
            fill = DATA_FILL
            font = DATA_FONT
            lc.alignment = LEFT_DATA
        lc.fill = fill
        lc.font = font

        # For each side (ACT, UW), emit $ and % formulas.
        for side_cols, category in ((cols_act, "ACT"), (cols_uw, "UW")):
            c_dollar = side_cols[0]
            c_pct    = side_cols[1]

            # $ formula
            raw = formula_for_dollars(tag, category, row_map)
            f_dollar = raw.replace("{COL}", c_dollar)
            cell = ws[f"{c_dollar}{r}"]
            cell.value = f_dollar
            cell.number_format = FMT_CURRENCY
            cell.alignment = RIGHT
            cell.fill = fill
            cell.border = BORDER_ALL
            cell.font = font

            # % of GPR formula
            if tag in PCT_DENOM_REVENUE_BLOCK:
                denom_row = row_map["ADJ_GPR"]
            else:
                denom_row = row_map["EFF_GROSS_REV"]
            f_pct = f'=IFERROR({c_dollar}{r}/{c_dollar}${denom_row},0)'
            cell = ws[f"{c_pct}{r}"]
            cell.value = f_pct
            cell.number_format = FMT_PCT
            cell.alignment = RIGHT
            cell.fill = fill
            cell.border = BORDER_ALL
            cell.font = font

        r += 1

    # ---------- column widths ------------------------------------------------
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    for col in cols_act + cols_uw:
        ws.column_dimensions[col].width = 16


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=False)
    if "_DealFinancial" not in wb.sheetnames:
        raise SystemExit("ERROR: sheet '_DealFinancial' not found.")
    if "Financials" not in wb.sheetnames:
        raise SystemExit("ERROR: sheet 'Financials' not found.")
    ws = wb["Financials"]
    build_table(ws)
    wb.save(WORKBOOK)
    print("Saved:", WORKBOOK)


if __name__ == "__main__":
    main()
